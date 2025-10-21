import tkinter as tk
from tkinter import filedialog, messagebox, Toplevel, Text, Scrollbar, RIGHT, Y, END
from tkinter import ttk
import threading
import pandas as pd
import re
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import traceback

# --- Thread-safe messagebox wrapper ---
_root_window = None

def _safe_messagebox(msg_type, title, message):
    """Thread-safe messagebox that schedules UI calls on main thread."""
    def show():
        if msg_type == 'error':
            messagebox.showerror(title, message)
        elif msg_type == 'info':
            messagebox.showinfo(title, message)
        elif msg_type == 'warning':
            messagebox.showwarning(title, message)
    
    if _root_window and threading.current_thread() != threading.main_thread():
        try:
            _root_window.after(0, show)
        except Exception:
            # If scheduling fails, just print to console
            print(f"{msg_type.upper()}: {title} - {message}")
    else:
        show()

def show_error(title, message):
    _safe_messagebox('error', title, message)

def show_info(title, message):
    _safe_messagebox('info', title, message)

def show_warning(title, message):
    _safe_messagebox('warning', title, message)

# --- Helper Functions ---

def clean_desk_name(val):
    val = str(val).strip()
    val = val.replace(',', '').replace('.', '').replace('_', '').replace(' ', '')
    val = val.capitalize()
    digits = re.findall(r'\d+', val)
    if len(digits) == 1 and len(digits[0]) == 3:
        padded = digits[0] + '0'
        val = re.sub(r'\b' + digits[0] + r'\b', padded, val)
    return val


def split_desk_id(val):
    val = str(val).strip()
    digits = re.findall(r'\d+', val)
    room = re.sub(r'\d+', '', val).strip().title()
    desk_number = digits[0] if digits else ''
    if val.lower() == 'blanks' or val == '':
        room = 'Blanks'
        desk_number = ''
    return pd.Series([room, desk_number])


def normalize_desk_series(series):
    """
    Vectorized normalization for a Desk_ID pandas Series:
    - convert to string
    - remove commas, dots, underscores, spaces
    - capitalize
    - if value ends with exactly 3 digits (optionally prefixed by letters), append a '0' to the digits
    - convert empty strings to 'Blanks'
    """
    s = series.fillna('').astype(str)

    # remove any non-alphanumeric characters (dots, commas, spaces, etc.)
    s = s.str.replace(r'[^A-Za-z0-9]+', '', regex=True)

    # capitalize (first letter upper, rest lower) for readability
    s = s.str.capitalize()

    # mask: whole string = letters (optional) followed by exactly 3 digits
    mask = s.str.match(r'^[A-Za-z]*\d{3}$')

    # append '0' to 3-digit numbers (on masked rows)
    if mask.any():
        # avoid regex backreference issues (e.g. '\\10') by using simple concatenation
        s_loc = s[mask] + '0'
        s = s.where(~mask, s_loc)

    # convert empty strings to 'Blanks'
    s = s.replace('', 'Blanks')

    return s


def normalize_serial_series(series):
    """Normalize serial numbers: remove non-alphanumeric, uppercase, strip."""
    s = series.fillna('').astype(str)
    s = s.str.strip()
    s = s.str.replace(r'[^A-Za-z0-9]+', '', regex=True)
    s = s.str.upper()
    s = s.replace('', pd.NA)
    return s


def preview_mismatches(dataframe):
    preview_window = Toplevel()
    preview_window.title("Mismatch Preview")

    text_area = Text(preview_window, wrap='word', width=100, height=30)
    scrollbar = Scrollbar(preview_window, command=text_area.yview)
    text_area.configure(yscrollcommand=scrollbar.set)

    for _, row in dataframe.iterrows():
        line = f"Room: {row['Room']}, Desk: {row['Desk_Number']}\nOnly in File 1: {row['Only_in_File1']}\nOnly in File 2: {row['Only_in_File2']}\n\n"
        text_area.insert(END, line)

    text_area.pack(side='left', fill='both', expand=True)
    scrollbar.pack(side=RIGHT, fill=Y)


def _derive_serial_from_skan(skan_val, skan2_val):
    """Given skan and skan2 values, derive a replacement serial per rules:
    - if skan/skan2 contains a 'V' (case-insensitive), return substring from 'V' to end
    - else if skan/skan2 contains digits starting with '6' (pattern r'6\\d+'), return 'V' + that digits
    - otherwise return None
    """
    for s in (skan_val, skan2_val):
        if s is None:
            continue
        s = str(s).strip()
        if not s:
            continue
        # look for 'V' or 'v'
        m = re.search(r'[Vv].*', s)
        if m:
            return m.group(0)
        # look for digits starting with 6
        m2 = re.search(r'6\d+', s)
        if m2:
            return 'V' + m2.group(0)
    return None


def _apply_skan_replacements(df, serial_columns, cb=None):
    """For each column in serial_columns, replace values starting with '0'
    by deriving from skan/skan2 when possible. Returns count of replacements and sample list.
    """
    replaced = 0
    samples = []
    if df is None or len(serial_columns) == 0:
        return replaced, samples
    if 'skan' not in df.columns and 'skan2' not in df.columns:
        return replaced, samples

    for col in serial_columns:
        if col not in df.columns:
            continue
        # Convert categorical to object to allow modifications
        if df[col].dtype.name == 'category':
            df[col] = df[col].astype(object)
        
        # create boolean mask for entries starting with '0'
        mask = df[col].astype(str).str.startswith('0', na=False)
        if not mask.any():
            continue
        idxs = df.index[mask]
        for i in idxs:
            skan_val = df.at[i, 'skan'] if 'skan' in df.columns else None
            skan2_val = df.at[i, 'skan2'] if 'skan2' in df.columns else None
            new_serial = _derive_serial_from_skan(skan_val, skan2_val)
            if new_serial:
                old = df.at[i, col]
                df.at[i, col] = new_serial
                replaced += 1
                if len(samples) < 5:
                    samples.append((col, old, new_serial))
    return replaced, samples


def save_formatted_excel(dataframe, output_path):
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Desk Mismatches")
        wb.active = ws
    else:
        ws.title = "Desk Mismatches"

    # Make a safe copy and convert pandas NA / NaN to empty strings
    df = dataframe.copy()
    try:
        df = df.fillna('')
    except Exception:
        # fallback: replace common NA markers
        df = df.replace({pd.NA: '', None: ''})

    # Move 'Dom' rows to the end
    if 'Room' in df.columns:
        dom_mask = df['Room'].astype(str).str.strip().str.lower() == 'dom'
        df_dom = df[dom_mask]
        df_rest = df[~dom_mask]
        df = pd.concat([df_rest, df_dom], ignore_index=True)

    # Add headers with bold and centered formatting (use actual dataframe columns)
    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Color for highlighting 'Dom' rows
    from openpyxl.styles import PatternFill
    dom_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # yellow

    # Append data rows safely
    for idx, row in enumerate(df.itertuples(index=False), start=2):
        safe_row = []
        for v in row:
            # convert pandas NA to empty string for Excel
            if v is pd.NA:
                safe_row.append('')
            else:
                safe_row.append(v)
        ws.append(safe_row)
        # Highlight Dom rows
        if 'Room' in df.columns and str(row[0]).strip().lower() == 'dom':
            for cell in ws[idx]:
                cell.fill = dom_fill

    # Apply alignment and wrapping to all data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = length + 2

    wb.save(output_path)


# --- Progress-aware comparison ---

def compare_excels(file1, file2, output_folder, progress_callback=None):
    """Compares two Excel files and calls progress_callback(percentage, message).

    The progress_callback is optional and should be called from the thread running
    compare_excels; the GUI wrapper will arrange for it to safely update the UI.
    """
    def cb(percent, msg):
        if progress_callback:
            progress_callback(percent, msg)

    # Validate file paths
    cb(0, 'Validating paths...')
    if not os.path.exists(file1):
        cb(0, f'File not found: {file1}')
        show_error("Error", f"File not found: {file1}")
        return
    if not os.path.exists(file2):
        cb(0, f'File not found: {file2}')
        show_error("Error", f"File not found: {file2}")
        return
    
    # Validate files are readable
    try:
        with open(file1, 'rb') as f:
            pass
    except PermissionError:
        cb(0, f'Permission denied: {file1}')
        show_error("Error", f"Cannot read file (permission denied):\n{file1}")
        return
    except Exception as e:
        cb(0, f'Cannot access file: {file1}')
        show_error("Error", f"Cannot access file:\n{file1}\n{e}")
        return
    
    try:
        with open(file2, 'rb') as f:
            pass
    except PermissionError:
        cb(0, f'Permission denied: {file2}')
        show_error("Error", f"Cannot read file (permission denied):\n{file2}")
        return
    except Exception as e:
        cb(0, f'Cannot access file: {file2}')
        show_error("Error", f"Cannot access file:\n{file2}\n{e}")
        return
    
    # Ensure output folder exists and is writable
    try:
        os.makedirs(output_folder, exist_ok=True)
        test_file = os.path.join(output_folder, '.test_write')
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
    except Exception as e:
        cb(0, f'Output folder not writable: {output_folder}')
        show_error("Error", f"Cannot write to output folder:\n{output_folder}\n{e}")
        return

    # Read Excel files in chunks with optimized memory usage
    cb(5, 'Reading Excel files (headers only)...')
    try:
        def pick_usecols(cols):
            usecols = []
            for c in cols:
                lc = str(c).strip().lower()
                # keep any serial columns starting with S.N
                if str(c).startswith('S.N'):
                    usecols.append(c)
                    continue
                # keep desk/place/room/type/skan columns
                if any(k in lc for k in ('desk', 'place', 'room', 'type', 'skan')):
                    usecols.append(c)
            # if nothing selected, return None to read all columns
            return usecols if usecols else None

        # Initialize header info
        cols1 = pd.read_excel(file1, nrows=0).columns.tolist()
        cols2 = pd.read_excel(file2, nrows=0).columns.tolist()
        use1 = pick_usecols(cols1)
        use2 = pick_usecols(cols2)
        
        cb(6, f'Columns to read - file1: {use1 or "ALL"} | file2: {use2 or "ALL"}')

        # Function to read Excel in chunks
        def read_excel_optimized(file_path, usecols):
            # Read in smaller chunks to reduce memory usage
            df = pd.read_excel(file_path, usecols=usecols)
            
            # Immediately filter MNTR if Type column exists to reduce memory
            if 'Type' in df.columns:
                df = df[df['Type'] == 'MNTR'].copy()
                df.reset_index(drop=True, inplace=True)
            
            # Note: Categorical conversion removed to avoid issues with later modifications
            
            return df

        cb(7, 'Reading and optimizing file 1...')
        df1 = read_excel_optimized(file1, use1)
        cb(8, 'Reading and optimizing file 2...')
        df2 = read_excel_optimized(file2, use2)
    except PermissionError as e:
        cb(0, 'Permission denied reading Excel files')
        show_error("Error", f"Permission denied. Please close the Excel files if they are open:\n{e}")
        return
    except Exception as e:
        cb(0, f'Failed to read Excel files: {e}')
        show_error("Error", f"Failed to read Excel files:\n{e}\n\nMake sure the files are valid Excel files (.xlsx or .xls)")
        return
    
    # Validate dataframes are not empty
    if df1 is None or df1.empty:
        cb(0, 'File 1 is empty or has no data')
        show_error("Error", f"File 1 contains no data:\n{file1}")
        return
    if df2 is None or df2.empty:
        cb(0, 'File 2 is empty or has no data')
        show_error("Error", f"File 2 contains no data:\n{file2}")
        return

    # Normalize possible Desk_ID-like column names to 'Desk_ID' for downstream code
    def ensure_desk_col(df):
        if 'Desk_ID' not in df.columns:
            for c in df.columns:
                if 'desk' in str(c).lower():
                    df = df.rename(columns={c: 'Desk_ID'})
                    break
        return df

    df1 = ensure_desk_col(df1)
    df2 = ensure_desk_col(df2)

    cb(20, 'Filtering MNTR rows...')
    try:
        cb(21, 'Filtering MNTR rows in file 1...')
        if 'Type' in df1.columns:
            df1 = df1[df1['Type'] == 'MNTR']
        cb(23, 'Filtering MNTR rows in file 2...')
        if 'Type' in df2.columns:
            df2 = df2[df2['Type'] == 'MNTR']
        cb(25, 'MNTR filtering complete')
    except Exception as e:
        cb(0, f'Error filtering MNTR rows: {e}')
        show_error("Error", f"Error filtering MNTR rows:\n{e}")
        return
    # Improved logic: do NOT forward-fill Desk_ID if only Place/Room is set.
    # Instead, if Desk_ID is blank and Place/Room is set, assign Desk_ID to Place/Room only.
    def fix_desk_ids(df):
        if 'Desk_ID' not in df.columns:
            df['Desk_ID'] = pd.Series([pd.NA] * len(df))
            return df
        # Try to find a Place/Room column (case-insensitive)
        place_col = None
        for col in df.columns:
            if col.strip().lower() in ('place', 'room'):
                place_col = col
                break
        # Replace empty Desk_ID with Place/Room if available, else leave blank
        desk_ids = df['Desk_ID'].replace(r'^\s*$', pd.NA, regex=True)
        if place_col:
            # Only fill Desk_ID with Place/Room if Desk_ID is blank and Place/Room is not blank
            mask = desk_ids.isna() & df[place_col].notna() & (df[place_col].astype(str).str.strip() != '')
            desk_ids = desk_ids.where(~mask, df[place_col])
        df['Desk_ID'] = desk_ids
        return df

    cb(26, 'Fixing Desk_IDs in file 1...')
    df1 = fix_desk_ids(df1)
    cb(28, 'Fixing Desk_IDs in file 2...')
    df2 = fix_desk_ids(df2)

    cb(30, 'Starting Desk ID normalization...')
    try:
        # Check if Desk_ID columns exist and show their sizes
        cb(31, f'File 1 shape: {df1.shape}, File 2 shape: {df2.shape}')
        
        # Normalize first file
        cb(32, 'Normalizing File 1 Desk IDs...')
        df1_desk_col = df1.get('Desk_ID', pd.Series(dtype=str))
        cb(33, f'File 1 Desk_ID count: {len(df1_desk_col)}')
        df1['Desk_ID'] = normalize_desk_series(df1_desk_col)
        
        # Normalize second file
        cb(34, 'Normalizing File 2 Desk IDs...')
        df2_desk_col = df2.get('Desk_ID', pd.Series(dtype=str))
        cb(35, f'File 2 Desk_ID count: {len(df2_desk_col)}')
        df2['Desk_ID'] = normalize_desk_series(df2_desk_col)
        
        # Show samples from both files
        try:
            sample_vals1 = ', '.join(df1['Desk_ID'].head(3).astype(str).tolist())
            sample_vals2 = ', '.join(df2['Desk_ID'].head(3).astype(str).tolist())
            cb(36, f'Samples - File1: {sample_vals1} | File2: {sample_vals2}')
        except Exception as e:
            cb(36, f'Could not show samples: {str(e)}')
            
        cb(37, 'Desk ID normalization completed successfully')
    except Exception as e:
        cb(30, f'Error during normalization: {str(e)}')
        show_error("Error", f"Failed during desk ID normalization:\n{str(e)}")
        return

    cb(40, 'Identifying serial columns...')
    serial_columns_file1 = [col for col in df1.columns if col.startswith('S.N')]
    serial_columns_file2 = [col for col in df2.columns if col.startswith('S.N')]
    
    # Warn if no serial columns found
    if not serial_columns_file1 and not serial_columns_file2:
        cb(0, 'No serial columns found in either file')
        show_error("Error", "No serial columns found!\n\nSerial columns must start with 'S.N' (e.g., S.N1, S.N2)")
        return
    
    if not serial_columns_file1:
        cb(41, f'Warning: No serial columns in file 1')
        show_warning("Warning", f"No serial columns found in file 1:\n{file1}\n\nAll serials will appear as 'Only in File 2'")
    
    if not serial_columns_file2:
        cb(41, f'Warning: No serial columns in file 2')
        show_warning("Warning", f"No serial columns found in file 2:\n{file2}\n\nAll serials will appear as 'Only in File 1'")

    # Apply skan/skan2-based replacements for serials that start with '0'
    cb(42, 'Fixing serials starting with 0 using skan/skan2...')
    replaced1, samples1 = _apply_skan_replacements(df1, serial_columns_file1)
    replaced2, samples2 = _apply_skan_replacements(df2, serial_columns_file2)
    if replaced1 or replaced2:
        sample_text = '; '.join([f"{c}:{o}->{n}" for c, o, n in (samples1 + samples2)])
        cb(44, f'Replaced {replaced1} in file1, {replaced2} in file2. Samples: {sample_text}')

    cb(45, 'Reshaping data...')
    # Handle empty serial columns gracefully
    if serial_columns_file1:
        df1_normalized = df1.melt(id_vars='Desk_ID', value_vars=serial_columns_file1, value_name='Serial_Number')
        df1_normalized = df1_normalized.dropna(subset=['Serial_Number'])
        # Free memory
        df1_normalized = df1_normalized[['Desk_ID', 'Serial_Number']].copy()
    else:
        df1_normalized = pd.DataFrame(columns=['Desk_ID', 'Serial_Number'])
    
    if serial_columns_file2:
        df2_normalized = df2.melt(id_vars='Desk_ID', value_vars=serial_columns_file2, value_name='Serial_Number')
        df2_normalized = df2_normalized.dropna(subset=['Serial_Number'])
        # Free memory
        df2_normalized = df2_normalized[['Desk_ID', 'Serial_Number']].copy()
    else:
        df2_normalized = pd.DataFrame(columns=['Desk_ID', 'Serial_Number'])

    # Split Desk_ID into Room and Desk_Number
    cb(55, 'Splitting Desk IDs...')
    df1_normalized[['Room', 'Desk_Number']] = df1_normalized['Desk_ID'].apply(split_desk_id)
    df2_normalized[['Room', 'Desk_Number']] = df2_normalized['Desk_ID'].apply(split_desk_id)

    # Normalize serials to avoid formatting mismatches
    cb(65, 'Normalizing serials...')
    df1_normalized['Serial_Number'] = normalize_serial_series(df1_normalized['Serial_Number'])
    df2_normalized['Serial_Number'] = normalize_serial_series(df2_normalized['Serial_Number'])

    # Diagnostic: same serial assigned to different desks across files
    merged_serials = pd.merge(df1_normalized[['Desk_ID', 'Serial_Number']].dropna(),
                              df2_normalized[['Desk_ID', 'Serial_Number']].dropna(),
                              on='Serial_Number', how='inner', suffixes=('_1', '_2'))
    diff_assign = merged_serials[merged_serials['Desk_ID_1'] != merged_serials['Desk_ID_2']]
    if not diff_assign.empty:
        sample = diff_assign[['Serial_Number', 'Desk_ID_1', 'Desk_ID_2']].drop_duplicates().head(5)
        sample_text = '; '.join(f"{r.Serial_Number}:{r.Desk_ID_1}!={r.Desk_ID_2}" for r in sample.itertuples())
        cb(67, f'Serial assigned to different desks across files (sample): {sample_text}')

    # Build serial -> desk maps to help fill missing desk info later
    cb(69, 'Building serial maps and sets...')
    # Build serial -> desk sets so we can detect ambiguous mappings.
    # Only use a serial->desk mapping later if the serial maps to exactly one non-blank desk.
    # Build serial -> desk maps using pandas groupby (only keep serials mapping to exactly one non-blank desk)
    def build_unique_serial_map(df_pairs):
        df_pairs = df_pairs.copy()
        df_pairs['Desk_ID'] = df_pairs['Desk_ID'].astype(str).str.strip()
        df_pairs = df_pairs[~df_pairs['Desk_ID'].isin(['', 'nan'])]
        df_pairs = df_pairs[~df_pairs['Desk_ID'].str.lower().eq('blanks')]
        grouped = df_pairs.groupby('Serial_Number')['Desk_ID'].nunique()
        unique_serials = grouped[grouped == 1].index
        # For those serials, grab the single Desk_ID (first value)
        single_map = df_pairs[df_pairs['Serial_Number'].isin(unique_serials)].drop_duplicates(subset=['Serial_Number'])
        return dict(zip(single_map['Serial_Number'], single_map['Desk_ID']))

    # Create pairs dataframes used for building serial->desk maps
    try:
        df1_pairs = df1_normalized[['Desk_ID', 'Serial_Number']].dropna().copy()
        df2_pairs = df2_normalized[['Desk_ID', 'Serial_Number']].dropna().copy()
        cb(76, f'Created pairs dataframes - File1: {len(df1_pairs)} rows, File2: {len(df2_pairs)} rows')
    except Exception:
        df1_pairs = pd.DataFrame(columns=['Desk_ID', 'Serial_Number'])
        df2_pairs = pd.DataFrame(columns=['Desk_ID', 'Serial_Number'])

    serial_to_desk_1 = build_unique_serial_map(df1_pairs)
    serial_to_desk_2 = build_unique_serial_map(df2_pairs)

    # Also keep multis for room-inference
    def build_multi_map(df_pairs):
        df_pairs = df_pairs.copy()
        df_pairs['Desk_ID'] = df_pairs['Desk_ID'].astype(str).str.strip()
        df_pairs = df_pairs[~df_pairs['Desk_ID'].isin(['', 'nan'])]
        df_pairs = df_pairs[~df_pairs['Desk_ID'].str.lower().eq('blanks')]
        multi = df_pairs.groupby('Serial_Number')['Desk_ID'].apply(lambda s: set(s.tolist())).to_dict()
        return multi

    serial_to_desk_1_multi = build_multi_map(df1_pairs)
    serial_to_desk_2_multi = build_multi_map(df2_pairs)

    # Helper: try to infer a common Room from multi mappings across both files.
    def _infer_room_from_serial(serial):
        union = set()
        if serial in serial_to_desk_1_multi:
            union |= set(serial_to_desk_1_multi[serial])
        if serial in serial_to_desk_2_multi:
            union |= set(serial_to_desk_2_multi[serial])
        # remove blank/Blanks
        union = {d for d in union if d and str(d).strip().lower() != 'blanks'}
        if not union:
            return None
        rooms = set()
        for d in union:
            room, _ = split_desk_id(d)
            if room:
                rooms.add(room)
        # if all mapped desks share the same room, return it
        if len(rooms) == 1:
            return next(iter(rooms))
        return None

    cb(75, 'Computing desk-centric mismatches...')

    def desk_serial_map(df_pairs):
        # Memory-optimized mapping function
        temp = df_pairs.copy()
        temp['Desk_ID'] = temp['Desk_ID'].astype(str).str.strip()
        temp['Serial_Number'] = temp['Serial_Number'].astype(str).str.strip()
        # Use dictionary comprehension for better memory efficiency
        return {desk: set(serials) for desk, serials in 
                temp.groupby('Desk_ID')['Serial_Number'].apply(list).items()}

    # Build optimized maps with memory management
    cb(76, 'Building desk->serial maps...')
    map1 = desk_serial_map(df1_pairs)
    map2 = desk_serial_map(df2_pairs)

    # Clean up to free memory
    del df1_pairs
    del df2_pairs

    # Get all unique desk IDs sorted as strings
    all_desks = list(set(map1.keys()) | set(map2.keys()))
    all_desks.sort(key=str)  # Sort using string comparison
    cb(77, f'Total desks to compare: {len(all_desks)}')

    # Initialize result containers with memory pre-allocation
    estimated_mismatches = min(len(all_desks) * 2, 1000)  # Conservative estimate
    only_in_1 = []
    only_in_2 = []
    
    # Process in optimized batches
    batch_size = 50
    total_batches = (len(all_desks) + batch_size - 1) // batch_size
    
    def process_batch(desks):
        """Process a batch of desks efficiently"""
        batch_in_1 = []
        batch_in_2 = []
        
        for desk in desks:
            # Get sets efficiently using dict.get with empty set default
            s1 = map1.get(desk, set())
            s2 = map2.get(desk, set())
            
            # Compute differences
            if s1 or s2:  # Only process if either set has items
                diff1 = s1 - s2
                diff2 = s2 - s1
                if diff1:
                    batch_in_1.extend((desk, serial) for serial in diff1)
                if diff2:
                    batch_in_2.extend((desk, serial) for serial in diff2)
        
        return batch_in_1, batch_in_2
    
    # Process all batches
    for batch_idx in range(total_batches):
        start_idx = batch_idx * batch_size
        end_idx = min((batch_idx + 1) * batch_size, len(all_desks))
        batch_desks = all_desks[start_idx:end_idx]
        
        # Process batch and extend results
        batch_in_1, batch_in_2 = process_batch(batch_desks)
        only_in_1.extend(batch_in_1)
        only_in_2.extend(batch_in_2)
        
        if batch_idx % 5 == 0:
            progress = int(76 + 14 * (batch_idx + 1) / total_batches)
            cb(progress, f'Comparing desks... (batch {batch_idx + 1}/{total_batches})')

    # Prepare DataFrame for output with memory optimization
    cb(85, 'Preparing output...')
    
    # Pre-allocate output structures
    total = len(only_in_1) + len(only_in_2)
    rows = []
    processed = 0
    
    # Free memory from objects no longer needed
    del map1, map2
    
    # Use garbage collection to ensure memory is freed
    import gc
    gc.collect()
    for desk_id, serial in only_in_1:
        # if desk_id is blank, try to find desk for this serial from the other file
        if not desk_id or str(desk_id).strip().lower() == 'blanks':
            # only use mapping from other file when it is uniquely mapped
            other_desk = serial_to_desk_2.get(serial)
            if other_desk:
                desk_to_use = other_desk
            else:
                # try to infer common room from all observed mappings across both files
                inferred_room = _infer_room_from_serial(serial)
                if inferred_room:
                    # use room name only; split_desk_id will place desk number empty
                    desk_to_use = inferred_room
                else:
                    desk_to_use = 'Unassigned'
        else:
            desk_to_use = desk_id
        room, desk_num = split_desk_id(desk_to_use)
        rows.append({'Room': room, 'Desk_Number': desk_num, 'Only_in_File1': serial, 'Only_in_File2': ''})
        processed += 1
        if processed % 100 == 0 and total > 0:
            progress = 85 + int(10 * processed / total)
            cb(min(progress, 94), f'Preparing output... ({processed}/{total})')
    for desk_id, serial in only_in_2:
        # if desk_id is blank, try to find desk for this serial from the other file
        if not desk_id or str(desk_id).strip().lower() == 'blanks':
            other_desk = serial_to_desk_1.get(serial)
            if other_desk:
                desk_to_use = other_desk
            else:
                inferred_room = _infer_room_from_serial(serial)
                if inferred_room:
                    desk_to_use = inferred_room
                else:
                    desk_to_use = 'Unassigned'
        else:
            desk_to_use = desk_id
        room, desk_num = split_desk_id(desk_to_use)
        rows.append({'Room': room, 'Desk_Number': desk_num, 'Only_in_File1': '', 'Only_in_File2': serial})
        processed += 1
        if processed % 100 == 0 and total > 0:
            progress = 85 + int(10 * processed / total)
            cb(min(progress, 94), f'Preparing output... ({processed}/{total})')

    # Memory-efficient aggregation of rows
    if not rows:
        result_df = pd.DataFrame(columns=['Room', 'Desk_Number', 'Only_in_File1', 'Only_in_File2'])
    else:
        # Create DataFrame with optimized memory usage
        df_rows = pd.DataFrame(rows, columns=['Room', 'Desk_Number', 'Only_in_File1', 'Only_in_File2'])
        
        def join_nonempty(series):
            # Optimized using pandas operations
            valid = series.dropna()
            if len(valid) == 0:
                return ''
            valid = valid.astype(str).str.strip()
            valid = valid[valid != '']
            # Use unique() to maintain order and remove duplicates efficiently
            unique_vals = valid.unique()
            return ', '.join(unique_vals)

        # Group and aggregate - removed categorical conversion to avoid type conflicts
        grouped = (df_rows.groupby(['Room', 'Desk_Number'])
                  .agg({
                      'Only_in_File1': join_nonempty,
                      'Only_in_File2': join_nonempty,
                  })
                  .reset_index()
                  .copy())  # Create a clean copy
        
        # Clean up intermediate data
        del df_rows
        gc.collect()

        # Ensure columns exist and order is consistent
        for col in ['Only_in_File1', 'Only_in_File2']:
            if col not in grouped.columns:
                grouped[col] = ''
        result_df = grouped[['Room', 'Desk_Number', 'Only_in_File1', 'Only_in_File2']].copy()
        
        # Remove serials that appear in BOTH columns (false mismatches from desk inference)
        rows_to_keep = []
        for idx in result_df.index:
            val1 = result_df.at[idx, 'Only_in_File1']
            val2 = result_df.at[idx, 'Only_in_File2']
            
            # Convert to string and split into sets
            str1 = str(val1) if pd.notna(val1) and str(val1).strip() else ''
            str2 = str(val2) if pd.notna(val2) and str(val2).strip() else ''
            
            if not str1 and not str2:
                continue  # Skip empty rows
            
            # Split by comma and strip whitespace
            serials_1 = set(s.strip() for s in str1.split(',') if s.strip())
            serials_2 = set(s.strip() for s in str2.split(',') if s.strip())
            
            # Find common serials (these are NOT mismatches - both files have them)
            common = serials_1 & serials_2
            
            if common:
                # Remove common serials from both sets
                serials_1 -= common
                serials_2 -= common
            
            # Only keep row if there are actual mismatches remaining
            if serials_1 or serials_2:
                result_df.at[idx, 'Only_in_File1'] = ', '.join(sorted(serials_1)) if serials_1 else ''
                result_df.at[idx, 'Only_in_File2'] = ', '.join(sorted(serials_2)) if serials_2 else ''
                rows_to_keep.append(idx)
        
        # Filter to only rows with actual mismatches
        if rows_to_keep:
            result_df = result_df.loc[rows_to_keep].reset_index(drop=True)
        else:
            result_df = pd.DataFrame(columns=['Room', 'Desk_Number', 'Only_in_File1', 'Only_in_File2'])

    # Save results
    cb(95, 'Saving results...')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_folder, f'desk_mismatches_{timestamp}.xlsx')
    try:
        save_formatted_excel(result_df, output_path)
    except PermissionError:
        cb(0, 'Cannot save - file may be open')
        show_error("Error", f"Cannot save results - the output file may be open in Excel.\n\nPlease close any open Excel files and try again.\n\nOutput path:\n{output_path}")
        return
    except Exception as e:
        cb(0, f'Failed to save results: {e}')
        show_error("Error", f"Failed to save results:\n{e}\n\nOutput path:\n{output_path}")
        return

    cb(100, f'Done. Saved to: {output_path}')
    show_info("Done", f"Comparison finished. Results saved to:\n{output_path}")


# --- GUI wrapper with progress bar ---

def start_comparison(root, file1, file2, output_folder, start_button):
    """Starts comparison on a background thread and shows a progress window."""
    progress_win = Toplevel(root)
    progress_win.title('Comparison Progress')
    progress_label = tk.Label(progress_win, text='Starting...')
    progress_label.pack(padx=10, pady=(10, 0))

    progressbar = ttk.Progressbar(progress_win, orient='horizontal', length=400, mode='determinate')
    progressbar.pack(padx=10, pady=10)

    percent_label = tk.Label(progress_win, text='0%')
    percent_label.pack(padx=10, pady=(0, 10))

    # Disable start button
    start_button.config(state='disabled')

    def progress_callback(percent, message):
        # This callback may be invoked from the worker thread; use after() to update UI
        def ui_update():
            progressbar['value'] = percent
            progress_label.config(text=message)
            percent_label.config(text=f"{percent}%")
            if percent >= 100:
                start_button.config(state='normal')
                # auto-close the progress window after short delay
                progress_win.after(800, progress_win.destroy)
        root.after(0, ui_update)

    def worker():
        try:
            compare_excels(file1, file2, output_folder, progress_callback=progress_callback)
        except Exception as e:
            # Ensure exceptions are reported back to the UI thread
            tb = traceback.format_exc()
            def report():
                progress_callback(0, f'Worker exception: {e}')
                show_error('Error in worker thread', f"{e}\n\nTraceback:\n{tb}")
                start_button.config(state='normal')
            root.after(0, report)

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()


# --- GUI ---

def build_demo_gui():
    global _root_window
    root = tk.Tk()
    _root_window = root
    root.title('Desk Comparator')

    frm = tk.Frame(root, padx=10, pady=10)
    frm.pack()

    tk.Label(frm, text='Plik 1:').grid(row=0, column=0, sticky='e')
    file1_var = tk.StringVar()
    tk.Entry(frm, textvariable=file1_var, width=60).grid(row=0, column=1)
    tk.Button(frm, text='Browse', command=lambda: file1_var.set(filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx;*.xls')]))).grid(row=0, column=2)

    tk.Label(frm, text='Plik 2:').grid(row=1, column=0, sticky='e')
    file2_var = tk.StringVar()
    tk.Entry(frm, textvariable=file2_var, width=60).grid(row=1, column=1)
    tk.Button(frm, text='Browse', command=lambda: file2_var.set(filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx;*.xls')]))).grid(row=1, column=2)

    tk.Label(frm, text='Folder wyjściowy:').grid(row=2, column=0, sticky='e')
    out_var = tk.StringVar()
    tk.Entry(frm, textvariable=out_var, width=60).grid(row=2, column=1)
    tk.Button(frm, text='Browse', command=lambda: out_var.set(filedialog.askdirectory())).grid(row=2, column=2)

    start_btn = tk.Button(frm, text='Rozpocznij porównanie', width=20, command=lambda: start_comparison(root, file1_var.get(), file2_var.get(), out_var.get() or os.getcwd(), start_btn))
    start_btn.grid(row=3, column=1, pady=10)

    root.mainloop()


if __name__ == '__main__':
    build_demo_gui()
